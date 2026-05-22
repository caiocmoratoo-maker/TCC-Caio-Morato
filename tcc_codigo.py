# pip install pandas numpy scipy matplotlib seaborn openpyxl xlrd

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import load_workbook
import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import seaborn as sns
import subprocess, warnings, os

warnings.filterwarnings("ignore")
sns.set_style("whitegrid")

PASTA_DADOS = "."
DIR_SAIDA = "resultados_tcc"
ESTADOS_COMM = ["PA", "MT", "MS", "GO", "MA", "TO", "MG"]
ESTADOS_IND = ["SP", "SC", "RS", "PR", "RJ", "AM"]
ANO_TRAT = 2005

COD_UF = {
    '11':'RO','12':'AC','13':'AM','14':'RR','15':'PA','16':'AP','17':'TO',
    '21':'MA','22':'PI','23':'CE','24':'RN','25':'PB','26':'PE','27':'AL',
    '28':'SE','29':'BA','31':'MG','32':'ES','33':'RJ','35':'SP',
    '41':'PR','42':'SC','43':'RS','50':'MS','51':'MT','52':'GO','53':'DF'
}
NOME_UF = {
    'Rondônia':'RO','Acre':'AC','Amazonas':'AM','Roraima':'RR','Pará':'PA',
    'Amapá':'AP','Tocantins':'TO','Maranhão':'MA','Piauí':'PI','Ceará':'CE',
    'Rio Grande do Norte':'RN','Paraíba':'PB','Pernambuco':'PE','Alagoas':'AL',
    'Sergipe':'SE','Bahia':'BA','Minas Gerais':'MG','Espírito Santo':'ES',
    'Rio de Janeiro':'RJ','São Paulo':'SP','Paraná':'PR','Santa Catarina':'SC',
    'Rio Grande do Sul':'RS','Mato Grosso do Sul':'MS','Mato Grosso':'MT',
    'Goiás':'GO','Distrito Federal':'DF'
}

# ── OLS com erros clusterizados ──────────────────────────────────────────────

def ols(X, y, cl=None, dof_adj=0):
    n, k = X.shape
    b = np.linalg.lstsq(X, y, rcond=None)[0]
    e = y - X @ b
    df = n - k - dof_adj
    if cl is not None:
        G = len(np.unique(cl))
        meat = sum(np.outer(s:=X[cl==c].T@e[cl==c], s) for c in np.unique(cl))
        V = np.linalg.inv(X.T@X) @ meat @ np.linalg.inv(X.T@X) * (G/(G-1)) * (n-1)/(n-k)
    else:
        V = (np.sum(e**2)/df) * np.linalg.inv(X.T@X)
    se = np.sqrt(np.diag(V))
    t = b/se
    p = 2*(1-stats.t.cdf(np.abs(t), df))
    r2 = 1 - np.sum(e**2)/np.sum((y-y.mean())**2)
    return {'b':b, 'se':se, 't':t, 'p':p, 'r2':r2, 'e':e, 'V':V, 'n':n, 'k':k}

def within(df, var, ent='estado'):
    return df[var] - df.groupby(ent)[var].transform('mean')

def within2(df, var, ent='estado', time='ano'):
    return df[var] - df.groupby(ent)[var].transform('mean') - df.groupby(time)[var].transform('mean') + df[var].mean()

def sig(p):
    return "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.1 else ""

def tabela(names, r, title=""):
    print(f"\n{title}\n"+"-"*70)
    print(f"{'Var':<18} {'Coef':>10} {'EP':>10} {'t':>8} {'p':>8}")
    print("-"*70)
    for i,nm in enumerate(names):
        print(f"{nm:<18} {r['b'][i]:>10.4f} {r['se'][i]:>10.4f} {r['t'][i]:>8.3f} {r['p'][i]:>8.4f} {sig(r['p'][i])}")
    print("-"*70)
    print(f"R²={r['r2']:.4f}  N={r['n']}")
    return r

# ── Carregamento de dados ────────────────────────────────────────────────────

def ler_ibge(path, sheet):
    wb = load_workbook(path, read_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True)); wb.close()
    years = [int(v) for v in rows[3] if v and str(v).isdigit()]
    recs = []
    for r in rows[4:]:
        vals = [v for v in r if v is not None]
        if len(vals)<4 or str(vals[0])=='1': continue
        s = COD_UF.get(str(vals[0]))
        if not s: continue
        for i,a in enumerate(years):
            if i<len(vals)-2 and isinstance(vals[i+2],(int,float)):
                recs.append({'estado':s,'ano':a,'valor':float(vals[i+2])})
    return pd.DataFrame(recs)

def ler_pop(path):
    wb = load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True)); wb.close()
    years = [int(v) for v in rows[3] if v and str(v).isdigit()]
    recs = []
    for r in rows[4:]:
        vals = [v for v in r if v is not None]
        if len(vals)<3: continue
        s = NOME_UF.get(str(vals[0]).strip())
        if not s: continue
        for i,a in enumerate(years):
            if i<len(vals)-1 and isinstance(vals[i+1],(int,float)):
                recs.append({'estado':s,'ano':a,'pop':float(vals[i+1])})
    df = pd.DataFrame(recs)
    out = []
    for uf in df['estado'].unique():
        sub = df[df['estado']==uf].set_index('ano').reindex(range(2002,2024))
        sub['pop'] = sub['pop'].interpolate(); sub['estado'] = uf
        out.append(sub.reset_index().rename(columns={'index':'ano'}))
    return pd.concat(out).dropna(subset=['pop'])

def ler_csv_bcb(path, skip_words):
    with open(path,"rb") as f: raw = f.read().decode("latin-1")
    recs = []
    for ln in raw.strip().split('\n'):
        if ';' not in ln or any(w in ln for w in skip_words): continue
        p = ln.strip().split(';')
        try:
            d,v = p[0].strip(), float(p[1].strip().replace(',','.'))
            if '/' in d:
                parts = d.split('/')
                if len(parts)==2:  # MM/YYYY
                    recs.append({'ano':int(parts[1]),'val':v})
                elif len(parts)==3:  # DD/MM/YYYY
                    recs.append({'ano':int(parts[2]),'val':v})
        except: pass
    return pd.DataFrame(recs)

def ler_trade(path):
    with open(path,"rb") as f: raw = f.read().decode("utf-8-sig")
    lines = [l for l in raw.strip().split('\n') if l.strip()]
    hdrs = lines[1].strip().split(';')
    years = [int(h) for h in hdrs[3:] if h.strip().isdigit()]
    recs = []
    for ln in lines[2:]:
        p = ln.strip().split(';')
        if len(p)<4 or len(p[0].strip())!=2: continue
        for i,a in enumerate(years):
            idx = 3+i
            if idx<len(p) and p[idx].strip():
                try: recs.append({'estado':p[0].strip(),'ano':a,'val':float(p[idx].strip())})
                except: pass
    return pd.DataFrame(recs)

# ── Construir painel ─────────────────────────────────────────────────────────

def construir_painel():
    P = PASTA_DADOS
    
    ind = ler_ibge(f"{P}/PIB_POR_ESTADO.xlsx","Tabela 27").rename(columns={'valor':'ind_pib'})
    pib = ler_ibge(f"{P}/PIB_POR_ESTADO.xlsx","Tabela 1").rename(columns={'valor':'pib_estado'})
    pop = ler_pop(f"{P}/Populaçao_estimada_por_estado.xlsx")
    
    cambio = ler_csv_bcb(f"{P}/TAXA_DE_CAMBIO_REAL.csv", ['Data','Índice'])
    cambio = cambio.groupby('ano')['val'].mean().reset_index().rename(columns={'val':'cambio_real'})
    
    comm_file = f"{P}/PREÇO_COMMMODITYS_INDEXADOS.xlsx"
    if not os.path.exists(comm_file):
        comm_file = f"{P}/PREÇO_COMMMODITYS_INDEXADOS.xls"
        if os.path.exists(comm_file):
            try: subprocess.run(['libreoffice','--headless','--convert-to','xlsx',comm_file,'--outdir',P],capture_output=True,timeout=30)
            except: pass
            xlsx = comm_file+'x'
            if os.path.exists(xlsx): comm_file = xlsx
    df_c = pd.read_excel(comm_file,header=0).iloc[2:].copy()
    df_c = df_c.rename(columns={df_c.columns[0]:'per',df_c.columns[1]:'v'})
    df_c['v'] = pd.to_numeric(df_c['v'],errors='coerce')
    df_c = df_c.dropna(subset=['v'])
    df_c['ano'] = df_c['per'].str.extract(r'(\d{4})').astype(int)
    comm = df_c.groupby('ano')['v'].mean().reset_index().rename(columns={'v':'commodities'})
    
    selic = ler_csv_bcb(f"{P}/TAXA_SELIC.csv", ['Data','Fonte','Taxa'])
    selic_a = selic.groupby('ano')['val'].mean().reset_index()
    selic_a['selic_aa'] = ((1+selic_a['val']/100)**252-1)*100
    
    ipca = ler_csv_bcb(f"{P}/IPCA.csv", ['Data','IPCA','Fonte'])
    ipca_a = ipca.groupby('ano').apply(lambda x: ((1+x['val']/100).prod()-1)*100).reset_index()
    ipca_a.columns = ['ano','ipca_aa']
    
    juros = selic_a[['ano','selic_aa']].merge(ipca_a, on='ano')
    juros['juros_reais'] = ((1+juros['selic_aa']/100)/(1+juros['ipca_aa']/100)-1)*100
    
    exp = ler_trade(f"{P}/EXPORTAÇÕES_POR_ESTADO.csv").rename(columns={'val':'exp'})
    imp = ler_trade(f"{P}/IMPORTAÇOES_POR_ESTADO.csv").rename(columns={'val':'imp'})
    trade = exp.merge(imp, on=['estado','ano'], how='outer').fillna(0)
    trade['xm'] = trade['exp'] + trade['imp']
    
    panel = ind.merge(pib, on=['estado','ano'], how='left')
    panel = panel.merge(pop[['estado','ano','pop']], on=['estado','ano'], how='left')
    panel['pib_pc'] = panel['pib_estado']/panel['pop']*1000
    panel = panel.merge(cambio, on='ano', how='left')
    panel = panel.merge(comm, on='ano', how='left')
    panel = panel.merge(juros[['ano','juros_reais']], on='ano', how='left')
    panel = panel.merge(trade[['estado','ano','xm']], on=['estado','ano'], how='left')
    panel['abertura'] = (panel['xm']/(panel['pib_estado']*1000))*100
    panel = panel.dropna(subset=['ind_pib','cambio_real','commodities','juros_reais'])
    panel = panel.sort_values(['estado','ano']).reset_index(drop=True)
    
    print(f"Painel: {len(panel)} obs, {panel['estado'].nunique()} UFs, {panel['ano'].min()}-{panel['ano'].max()}")
    return panel

# ── Estimações ───────────────────────────────────────────────────────────────

def estimar(panel):
    V = ['cambio_real','commodities','juros_reais','abertura','pib_pc']
    N_ent = panel['estado'].nunique()
    cl = panel['estado'].values
    y = panel['ind_pib'].values

    # Descritivas
    print("\n" + "="*60 + "\nESTATÍSTICAS DESCRITIVAS\n" + "="*60)
    print(panel[['ind_pib']+V].describe().round(4).to_string())

    # Pooled OLS
    X_pool = np.column_stack([np.ones(len(panel))]+[panel[v].values for v in V])
    pooled = tabela(['const']+V, ols(X_pool, y, cl), "POOLED OLS")

    # FE Two-Way
    df = panel.copy()
    for v_ in ['ind_pib']+V: df[f'{v_}_w'] = within2(df, v_)
    y_fe2 = df['ind_pib_w'].values
    X_fe2 = np.column_stack([df[f'{v}_w'].values for v in V])
    fe2 = tabela(V, ols(X_fe2, y_fe2, cl, N_ent+panel['ano'].nunique()-2), "FE TWO-WAY")

    # FE Entity
    df1 = panel.copy()
    for v_ in ['ind_pib']+V: df1[f'{v_}_w'] = within(df1, v_)
    y_fe = df1['ind_pib_w'].values
    X_fe = np.column_stack([df1[f'{v}_w'].values for v in V])
    fe = tabela(V, ols(X_fe, y_fe, cl, N_ent-1), "FE ENTITY")
    ss_w = np.sum((y - panel.groupby('estado')['ind_pib'].transform('mean').values)**2)
    fe['r2'] = 1 - np.sum(fe['e']**2)/ss_w
    print(f"R² within corrigido: {fe['r2']:.4f}")

    # RE (FGLS)
    Ti = panel.groupby('estado').size().values[0]
    s2w = np.sum(fe['e']**2)/(len(panel)-N_ent-len(V))
    means = panel.groupby('estado')[V+['ind_pib']].mean()
    Xb = np.column_stack([np.ones(len(means))]+[means[v].values for v in V])
    eb = means['ind_pib'].values - Xb @ np.linalg.lstsq(Xb, means['ind_pib'].values, rcond=None)[0]
    s2a = max(0, np.var(eb)-s2w/Ti)
    theta = 1-np.sqrt(s2w/(s2w+Ti*s2a))
    df_re = panel.copy()
    for v_ in ['ind_pib']+V:
        df_re[f'{v_}_re'] = df_re[v_] - theta*df_re.groupby('estado')[v_].transform('mean')
    X_re = np.column_stack([np.ones(len(df_re))*(1-theta)]+[df_re[f'{v}_re'].values for v in V])
    re = tabela(['const']+V, ols(X_re, df_re['ind_pib_re'].values, cl), f"RE (θ={theta:.4f})")

    # Hausman
    print("\n── HAUSMAN ──")
    d = fe['b'] - re['b'][1:]
    Vd = fe['V'] - re['V'][1:,1:]
    try:
        H = float(d.T @ np.linalg.inv(Vd) @ d)
        pv = 1-stats.chi2.cdf(H, len(V))
        print(f"χ²={H:.4f}, df={len(V)}, p={pv:.6f} → {'FE' if pv<0.05 else 'RE aceitável'}")
    except: print("Matriz singular → usar FE")

    # Interação CAMBIO × COMM
    panel['cxc'] = panel['cambio_real']*panel['commodities']
    Vi = V+['cxc']
    di = panel.copy()
    for v_ in ['ind_pib']+Vi: di[f'{v_}_w'] = within(di, v_)
    Xi = np.column_stack([di[f'{v}_w'].values for v in Vi])
    inter = tabela(Vi, ols(Xi, di['ind_pib_w'].values, cl, N_ent-1), "FE + INTERAÇÃO")
    ix = Vi.index('cxc')
    print(f"\n→ Interação: β={inter['b'][ix]:.6f}, p={inter['p'][ix]:.4f} {'✓ Doença Holandesa' if inter['p'][ix]<0.05 else '✗ Não significativo'}")

    # DiD
    print("\n" + "="*60 + "\nDIFERENÇAS-EM-DIFERENÇAS\n" + "="*60)
    dd = panel[panel['estado'].isin(ESTADOS_COMM+ESTADOS_IND)].copy()
    dd['T'] = dd['estado'].isin(ESTADOS_COMM).astype(int)
    dd['P'] = (dd['ano']>=ANO_TRAT).astype(int)
    dd['did'] = dd['T']*dd['P']
    ctrl = ['cambio_real','juros_reais','abertura','pib_pc']

    # Simples
    Xs = np.column_stack([np.ones(len(dd)), dd['T'].values, dd['P'].values, dd['did'].values])
    tabela(['const','trat','pos','DiD'], ols(Xs, dd['ind_pib'].values, dd['estado'].values), "DiD SIMPLES")

    # Com controles
    Xc = np.column_stack([np.ones(len(dd)), dd['T'].values, dd['P'].values, dd['did'].values]+[dd[c].values for c in ctrl])
    did_c = tabela(['const','trat','pos','DiD']+ctrl, ols(Xc, dd['ind_pib'].values, dd['estado'].values), "DiD + CONTROLES")

    # TWFE
    dd2 = dd.copy()
    tv = ['did']+ctrl
    for v_ in ['ind_pib']+tv: dd2[f'{v_}_w'] = within2(dd2, v_)
    Xt = np.column_stack([dd2[f'{v}_w'].values for v in tv])
    tabela(tv, ols(Xt, dd2['ind_pib_w'].values, dd2['estado'].values, dd2['estado'].nunique()+dd2['ano'].nunique()-2), "DiD TWFE")

    # Médias
    print("\nMédias:")
    for g,lb in [(1,"Trat"),(0,"Ctrl")]:
        pre = dd[(dd['T']==g)&(dd['P']==0)]['ind_pib'].mean()
        pos = dd[(dd['T']==g)&(dd['P']==1)]['ind_pib'].mean()
        print(f"  {lb}: pré={pre:.2f} pós={pos:.2f} Δ={pos-pre:+.2f}")

    # Robustez
    print("\n" + "="*60 + "\nROBUSTEZ\n" + "="*60)
    for nome, vs in {"s/ comm":['cambio_real','juros_reais','abertura','pib_pc'],
                     "s/ câmbio":['commodities','juros_reais','abertura','pib_pc'],
                     "só macro":['cambio_real','juros_reais','commodities'],
                     "abert+pib":['abertura','pib_pc']}.items():
        dr = panel.copy()
        for v_ in ['ind_pib']+vs: dr[f'{v_}_w'] = within(dr, v_)
        Xr = np.column_stack([dr[f'{v}_w'].values for v in vs])
        tabela(vs, ols(Xr, dr['ind_pib_w'].values, cl, N_ent-1), f"Robustez: {nome}")

    # Resumo
    print("\n" + "="*60 + f"\n{'Var':<18} {'Pooled':>12} {'FE':>12} {'FE+Inter':>12}\n" + "-"*55)
    for v_ in V+['cxc']:
        vals = []
        vals.append(f"{pooled['b'][V.index(v_)+1]:.4f}{sig(pooled['p'][V.index(v_)+1])}" if v_ in V else "—")
        vals.append(f"{fe['b'][V.index(v_)]:.4f}{sig(fe['p'][V.index(v_)])}" if v_ in V else "—")
        vals.append(f"{inter['b'][Vi.index(v_)]:.4f}{sig(inter['p'][Vi.index(v_)])}" if v_ in Vi else "—")
        print(f"{v_:<18} {vals[0]:>12} {vals[1]:>12} {vals[2]:>12}")
    print(f"{'R²':<18} {pooled['r2']:>12.4f} {fe['r2']:>12.4f} {inter['r2']:>12.4f}")

    return {'pooled':pooled,'fe':fe,'inter':inter,'did':did_c}

# ── Gráficos ─────────────────────────────────────────────────────────────────

def graficos(panel):
    D = DIR_SAIDA

    # 1. Evolução DiD
    fig,ax = plt.subplots(figsize=(11,6))
    for lb,est,c,m in [("Commodity",ESTADOS_COMM,"#e74c3c","o"),("Industrial",ESTADOS_IND,"#2980b9","s")]:
        med = panel[panel['estado'].isin(est)].groupby('ano')['ind_pib'].mean()
        ax.plot(med.index, med.values, marker=m, label=lb, color=c, lw=2.5, ms=6)
    ax.axvline(ANO_TRAT, color="gray", ls="--", alpha=.7, label=f"Boom ({ANO_TRAT})")
    ax.set(xlabel="Ano", ylabel="Ind/VAB (%)", title="Evolução Industrial: Commodity vs Industrial")
    ax.legend(); ax.set_xticks(range(2002,2022,2)); plt.tight_layout()
    plt.savefig(f"{D}/01_evolucao_did.png",dpi=150); plt.close()

    # 2. Tendências paralelas
    fig,ax = plt.subplots(figsize=(11,6))
    for lb,est,c,m in [("Commodity",ESTADOS_COMM,"#e74c3c","o"),("Industrial",ESTADOS_IND,"#2980b9","s")]:
        med = panel[panel['estado'].isin(est)].groupby('ano')['ind_pib'].mean()
        base = med.loc[med.index<ANO_TRAT].iloc[-1]
        ax.plot(med.index, (med/base)*100, marker=m, label=lb, color=c, lw=2.5, ms=6)
    ax.axvline(ANO_TRAT, color="gray", ls="--", alpha=.7)
    ax.axhline(100, color="gray", ls=":", alpha=.5)
    ax.set(xlabel="Ano", ylabel=f"Índice ({ANO_TRAT-1}=100)", title="Tendências Paralelas")
    ax.legend(); ax.set_xticks(range(2002,2022,2)); plt.tight_layout()
    plt.savefig(f"{D}/02_tendencias_paralelas.png",dpi=150); plt.close()

    # 3. Indústria vs Commodities
    fig,ax1 = plt.subplots(figsize=(11,6))
    ma = panel.groupby('ano')[['ind_pib','commodities']].mean()
    l1 = ax1.plot(ma.index, ma['ind_pib'], 'b-o', label='Ind/VAB (%)', lw=2.5, ms=6)
    ax1.set(xlabel="Ano", ylabel="Ind (%)")
    ax2 = ax1.twinx()
    l2 = ax2.plot(ma.index, ma['commodities'], 'r--s', label='Commodities', lw=2, ms=5, alpha=.8)
    ax2.set_ylabel("Commodities (2016=100)")
    ax1.legend(l1+l2, [x.get_label() for x in l1+l2], loc="upper right")
    ax1.set_title("Indústria vs Commodities"); ax1.set_xticks(range(2002,2022,2))
    plt.tight_layout(); plt.savefig(f"{D}/03_industria_vs_commodities.png",dpi=150); plt.close()

    # 4. Scatter câmbio vs indústria
    fig,ax = plt.subplots(figsize=(9,6))
    for s in panel['estado'].unique():
        sub = panel[panel['estado']==s]
        c = '#e74c3c' if s in ESTADOS_COMM else '#2980b9' if s in ESTADOS_IND else 'gray'
        a = 0.3 if c!='gray' else 0.15
        ax.scatter(sub['cambio_real'], sub['ind_pib'], alpha=a, s=15 if c!='gray' else 10, c=c)
    z = np.polyfit(panel['cambio_real'], panel['ind_pib'], 1)
    xl = np.linspace(panel['cambio_real'].min(), panel['cambio_real'].max(), 100)
    ax.plot(xl, np.poly1d(z)(xl), "k-", lw=2)
    ax.legend(handles=[Patch(fc='#e74c3c',alpha=.5,label='Commodity'),Patch(fc='#2980b9',alpha=.5,label='Industrial'),Patch(fc='gray',alpha=.3,label='Outros')])
    ax.set(xlabel="Câmbio Real", ylabel="Ind/VAB (%)", title="Câmbio vs Indústria")
    plt.tight_layout(); plt.savefig(f"{D}/04_scatter_cambio.png",dpi=150); plt.close()

    # 5. Correlação
    fig,ax = plt.subplots(figsize=(8,6))
    vs = ['ind_pib','cambio_real','commodities','juros_reais','abertura','pib_pc']
    lb = ['Ind','Câmbio','Comm','Juros','Abert','PIBpc']
    cr = panel[vs].corr(); cr.index=lb; cr.columns=lb
    sns.heatmap(cr, annot=True, fmt=".2f", cmap="RdBu_r", center=0, vmin=-1, vmax=1, ax=ax, square=True)
    ax.set_title("Correlação"); plt.tight_layout()
    plt.savefig(f"{D}/05_correlacao.png",dpi=150); plt.close()

    # 6. Séries macro
    fig,axes = plt.subplots(3,1,figsize=(11,10),sharex=True)
    ma = panel.groupby('ano')[['cambio_real','commodities','juros_reais']].mean()
    for ax,col,c,lb in zip(axes,['cambio_real','commodities','juros_reais'],['g','r','m'],['Câmbio','Commodities','Juros (%)']):
        ax.plot(ma.index, ma[col], f'{c}-o', lw=2, ms=5); ax.set_ylabel(lb)
        ax.axvline(ANO_TRAT, color="gray", ls="--", alpha=.5)
    axes[0].set_title("Variáveis Macro"); axes[2].set_xlabel("Ano")
    for ax in axes: ax.set_xticks(range(2002,2022,2))
    plt.tight_layout(); plt.savefig(f"{D}/06_macro.png",dpi=150); plt.close()

    # 7. Heatmap estados
    pv = panel.pivot_table(values='ind_pib', index='estado', columns='ano')
    pv = pv.loc[pv.mean(axis=1).sort_values(ascending=False).index]
    fig,ax = plt.subplots(figsize=(14,8))
    sns.heatmap(pv, cmap="YlOrRd_r", ax=ax, linewidths=.3, cbar_kws={'label':'Ind (%)'})
    ax.set_title("Ind/VAB por UF e Ano"); plt.tight_layout()
    plt.savefig(f"{D}/07_heatmap.png",dpi=150); plt.close()

    # 8. DiD com IC
    fig,ax = plt.subplots(figsize=(11,6))
    for lb,est,c in [("Commodity",ESTADOS_COMM,"#e74c3c"),("Industrial",ESTADOS_IND,"#2980b9")]:
        sub = panel[panel['estado'].isin(est)]
        st = sub.groupby('ano')['ind_pib'].agg(['mean','std','count'])
        st['se'] = st['std']/np.sqrt(st['count'])
        ax.plot(st.index, st['mean'], '-o', color=c, label=lb, lw=2)
        ax.fill_between(st.index, st['mean']-1.96*st['se'], st['mean']+1.96*st['se'], alpha=.2, color=c)
    ax.axvline(ANO_TRAT, color="gray", ls="--", alpha=.7, label=f"Trat ({ANO_TRAT})")
    ax.set(xlabel="Ano", ylabel="Ind (%)", title="DiD: Médias ± IC 95%")
    ax.legend(); ax.set_xticks(range(2002,2022,2)); plt.tight_layout()
    plt.savefig(f"{D}/08_did_ic.png",dpi=150); plt.close()

    print(f"✓ 8 gráficos salvos em {D}/")

# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    os.makedirs(DIR_SAIDA, exist_ok=True)
    panel = construir_painel()
    panel.to_csv(f"{DIR_SAIDA}/painel.csv", index=False)
    
    with open(f"{DIR_SAIDA}/resultados.txt","w",encoding="utf-8") as f:
        import sys
        old = sys.stdout
        sys.stdout = type('Tee',(object,),{'write':lambda s,x:(old.write(x),f.write(x)),'flush':lambda s:(old.flush(),f.flush())})()
        estimar(panel)
        sys.stdout = old
    
    graficos(panel)
    print(f"\n✓ Tudo salvo em {DIR_SAIDA}/")
